"""
SEKOLAH SCRAPER - SELENIUM COMPLETE VERSION
============================================
Scraping data sekolah dari sekolah.data.kemendikdasmen.go.id
Menggunakan Selenium dengan headless browser
"""

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from bs4 import BeautifulSoup
import pandas as pd
import time
import json
import os
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


class SekolahScraperSelenium:
    """
    Selenium-based scraper untuk data sekolah
    """

    def __init__(self, max_workers=2, headless=True, debug=False):
        """
        Parameters:
        - max_workers: jumlah browser instances untuk parallel scraping
        - headless: True untuk headless browser
        - debug: True untuk verbose logging
        """
        self.max_workers = max_workers
        self.headless = headless
        self.debug = debug

        self.base_url = "https://sekolah.data.kemendikdasmen.go.id"
        self.checkpoint_file = "checkpoint.json"
        self.temp_data_file = "temp_data.json"
        self.batch_size = 50  # simpan batch setiap 50 data
        self.batch_counter = 0
        
        self.lock = threading.Lock()
        self.all_data = []

        print("\n" + "="*70)
        print("  SEKOLAH SCRAPER - SELENIUM COMPLETE VERSION")
        print("="*70)
        print(f"  Workers: {max_workers}")
        print(f"  Headless: {headless}")
        print(f"  Debug: {debug}")
        print("="*70 + "\n")

    def create_driver(self):
        """Buat instance Selenium WebDriver"""
        try:
            chrome_options = Options()
            
            if self.headless:
                chrome_options.add_argument('--headless=new')
            
            chrome_options.add_argument('--no-sandbox')
            chrome_options.add_argument('--disable-dev-shm-usage')
            chrome_options.add_argument('--disable-blink-features=AutomationControlled')
            chrome_options.add_argument('--disable-gpu')
            chrome_options.add_argument('--window-size=1920,1080')
            chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
            
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)

            driver = webdriver.Chrome(options=chrome_options)
            driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            
            return driver

        except Exception as e:
            print(f"[ERROR] Gagal membuat driver: {e}")
            return None

    def get_total_schools(self):
        """Ambil total jumlah sekolah dari halaman pertama"""
        driver = self.create_driver()
        if not driver:
            return 0

        try:
            print("→ Mengambil jumlah total sekolah...")
            driver.get(f"{self.base_url}/sekolah")
            time.sleep(5)

            # coba ambil dari pagination atau count di halaman
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            
            # hitung dari artikel yang ada di halaman pertama
            articles = soup.find_all('article')
            
            if articles:
                print(f"✓ Halaman berhasil dimuat dengan {len(articles)} sekolah per halaman")
                # karena tidak ada total count yang jelas, scrape sampai habis
                return -1  # flag untuk scrape semua
            else:
                print("✗ Tidak ada sekolah ditemukan di halaman pertama")
                return 0

        except Exception as e:
            print(f"✗ Error mengambil total: {e}")
            return 0
        finally:
            driver.quit()

    def extract_school_list(self, driver):
        """Ekstrak daftar sekolah dari halaman list"""
        try:
            time.sleep(3)
            soup = BeautifulSoup(driver.page_source, 'html.parser')

            schools = []
            articles = soup.find_all('article')

            for article in articles:
                try:
                    # ekstrak NPSN
                    npsn = ''
                    npsn_elem = article.find(string=lambda x: x and 'NPSN' in str(x))
                    if npsn_elem:
                        npsn = str(npsn_elem).replace('NPSN', '').replace(':', '').strip()

                    # ekstrak nama sekolah
                    nama = ''
                    nama_elem = article.find('h3')
                    if nama_elem:
                        nama = nama_elem.get_text(strip=True)

                    # ekstrak status
                    status = ''
                    status_elem = article.find('div', class_=lambda x: x and 'font-semibold' in str(x))
                    if status_elem:
                        status = status_elem.get_text(strip=True)

                    # ekstrak alamat
                    alamat = ''
                    alamat_elem = article.find('div', class_=lambda x: x and 'line-clamp' in str(x))
                    if alamat_elem:
                        alamat = alamat_elem.get_text(strip=True)

                    if npsn:
                        schools.append({
                            'npsn': npsn,
                            'nama_sekolah': nama,
                            'status_sekolah': status,
                            'alamat_jalan': alamat,
                            'url': f"{self.base_url}/profil-sekolah/{npsn}"
                        })

                except Exception as e:
                    if self.debug:
                        print(f"    ⚠ Error parse artikel: {e}")
                    continue

            return schools

        except Exception as e:
            print(f"  ✗ Error ekstrak list: {e}")
            return []

    def scrape_school_detail(self, npsn, url):
        """Scrape detail sekolah dari halaman profil"""
        driver = self.create_driver()
        if not driver:
            return None

        try:
            driver.get(url)
            time.sleep(4)

            soup = BeautifulSoup(driver.page_source, 'html.parser')

            data = {
                'npsn': npsn,
                'url': url,
                'nama_sekolah': '-',
                'alamat_sekolah': '-',
                'url_sk_operasional': '-',
                'akreditasi': '-',
                'kepala_sekolah': '-',
                'telepon': '-',
                'website': '-',
                'status_sekolah': '-',
                'bentuk_pendidikan': '-',
                'operator': '-',
                'email': '-',
                'yayasan': '-',
                # statistik
                'guru': '-',
                'siswa_laki': '-',
                'siswa_perempuan': '-',
                'rombongan_belajar': '-',
                'daya_tampung': '-',
                'ruang_kelas': '-',
                'laboratorium': '-',
                'perpustakaan': '-',
                # kurikulum & utilitas
                'kurikulum': '-',
                'penyelenggaraan': '-',
                'akses_internet': '-',
                'sumber_listrik': '-',
                'daya_listrik': '-',
                'luas_tanah': '-',
                # proses pembelajaran
                'rasio_siswa_rombel': '-',
                'rasio_rombel_ruang_kelas': '-',
                'rasio_siswa_guru': '-',
                'persen_guru_kualifikasi': '-',
                'persen_guru_sertifikasi': '-',
                'persen_guru_pns': '-',
                'persen_ruang_kelas_layak': '-',
                # alamat
                'lintang': '-',
                'bujur': '-',
                'link_google_map': '-'
            }

            # ekstrak nama sekolah dari header
            h1 = soup.find('h1')
            if h1:
                data['nama_sekolah'] = h1.get_text(strip=True)

            # ekstrak data dari card profil umum
            profile_cards = soup.find_all('div', class_=lambda x: x and 'rounded' in str(x))
            
            for card in profile_cards:
                text = card.get_text()
                
                # alamat
                if 'Alamat' in text:
                    alamat_div = card.find('div', class_=lambda x: x and 'text-slate-600' in str(x))
                    if alamat_div:
                        data['alamat_sekolah'] = alamat_div.get_text(strip=True)
                
                # SK operasional
                if 'SK Operasional' in text:
                    link = card.find('a', href=True)
                    if link:
                        data['url_sk_operasional'] = link['href']
                
                # akreditasi
                if 'Akreditasi' in text:
                    akred_elem = card.find('div', class_=lambda x: x and 'font-semibold' in str(x))
                    if akred_elem:
                        data['akreditasi'] = akred_elem.get_text(strip=True)
                
                # kepala sekolah
                if 'Kepala Sekolah' in text:
                    kepala_elem = card.find('div', class_=lambda x: x and 'font-semibold' in str(x))
                    if kepala_elem:
                        data['kepala_sekolah'] = kepala_elem.get_text(strip=True)
                
                # telepon
                if 'Telepon' in text or 'No. Telp' in text:
                    telp_elem = card.find('a', href=lambda x: x and 'tel:' in str(x))
                    if telp_elem:
                        data['telepon'] = telp_elem.get_text(strip=True)
                
                # website
                if 'Website' in text:
                    web_elem = card.find('a', href=True)
                    if web_elem and 'http' in web_elem['href']:
                        data['website'] = web_elem['href']
                
                # status
                if 'Status' in text and 'Status Sekolah' not in data or data['status_sekolah'] == '-':
                    status_elem = card.find('div', class_=lambda x: x and 'font-semibold' in str(x))
                    if status_elem:
                        data['status_sekolah'] = status_elem.get_text(strip=True)
                
                # bentuk pendidikan
                if 'Bentuk Pendidikan' in text:
                    bentuk_elem = card.find('div', class_=lambda x: x and 'font-semibold' in str(x))
                    if bentuk_elem:
                        data['bentuk_pendidikan'] = bentuk_elem.get_text(strip=True)
                
                # operator
                if 'Operator' in text:
                    op_elem = card.find('div', class_=lambda x: x and 'font-semibold' in str(x))
                    if op_elem:
                        data['operator'] = op_elem.get_text(strip=True)
                
                # email
                if 'Email' in text:
                    email_elem = card.find('a', href=lambda x: x and 'mailto:' in str(x))
                    if email_elem:
                        data['email'] = email_elem.get_text(strip=True)
                
                # yayasan
                if 'Yayasan' in text:
                    yayasan_elem = card.find('div', class_=lambda x: x and 'font-semibold' in str(x))
                    if yayasan_elem:
                        data['yayasan'] = yayasan_elem.get_text(strip=True)

            # ekstrak statistik sekolah
            stat_items = soup.find_all('div', class_=lambda x: x and 'text-2xl' in str(x))
            stat_labels = soup.find_all('div', class_=lambda x: x and 'text-slate-500' in str(x))
            
            for i, (stat, label) in enumerate(zip(stat_items, stat_labels)):
                label_text = label.get_text(strip=True).lower()
                value = stat.get_text(strip=True)
                
                if 'guru' in label_text:
                    data['guru'] = value
                elif 'laki-laki' in label_text:
                    data['siswa_laki'] = value
                elif 'perempuan' in label_text:
                    data['siswa_perempuan'] = value
                elif 'rombongan' in label_text or 'rombel' in label_text:
                    data['rombongan_belajar'] = value
                elif 'daya tampung' in label_text:
                    data['daya_tampung'] = value
                elif 'ruang kelas' in label_text:
                    data['ruang_kelas'] = value
                elif 'laboratorium' in label_text:
                    data['laboratorium'] = value
                elif 'perpustakaan' in label_text:
                    data['perpustakaan'] = value

            # ekstrak kurikulum & utilitas
            all_text = soup.get_text()
            
            if 'Kurikulum' in all_text:
                kurikulum_match = soup.find(string=lambda x: x and 'Kurikulum' in str(x))
                if kurikulum_match:
                    parent = kurikulum_match.find_parent()
                    if parent:
                        next_div = parent.find_next('div', class_=lambda x: x and 'font-semibold' in str(x))
                        if next_div:
                            data['kurikulum'] = next_div.get_text(strip=True)

            # ekstrak koordinat dan link Google Maps
            if 'Lintang' in all_text:
                lintang_elem = soup.find(string=lambda x: x and 'Lintang' in str(x))
                if lintang_elem:
                    parent = lintang_elem.find_parent()
                    if parent:
                        coord_div = parent.find_next('div')
                        if coord_div:
                            data['lintang'] = coord_div.get_text(strip=True)
            
            if 'Bujur' in all_text:
                bujur_elem = soup.find(string=lambda x: x and 'Bujur' in str(x))
                if bujur_elem:
                    parent = bujur_elem.find_parent()
                    if parent:
                        coord_div = parent.find_next('div')
                        if coord_div:
                            data['bujur'] = coord_div.get_text(strip=True)
            
            # Google Maps link
            maps_link = soup.find('a', href=lambda x: x and 'maps.google.com' in str(x))
            if maps_link:
                data['link_google_map'] = maps_link['href']

            return data

        except Exception as e:
            if self.debug:
                print(f"  ✗ Error scrape detail untuk {npsn}: {e}")
            return None
        finally:
            driver.quit()

    def scrape_page(self, page_num):
        """Scrape satu halaman"""
        driver = self.create_driver()
        if not driver:
            return []

        try:
            url = f"{self.base_url}/sekolah"
            driver.get(url)
            time.sleep(5)

            # navigasi ke halaman yang diminta
            for i in range(1, page_num):
                try:
                    next_btn = driver.find_element(By.XPATH, "//button[contains(@class, 'p-paginator-next')]")
                    next_btn.click()
                    time.sleep(3)
                except:
                    print(f"  ✗ Tidak bisa navigasi ke halaman {page_num}")
                    return []

            # ekstrak school list
            schools = self.extract_school_list(driver)
            
            return schools

        except Exception as e:
            print(f"  ✗ Error halaman {page_num}: {e}")
            return []
        finally:
            driver.quit()

    def process_school(self, school_info):
        """Proses single school - ambil detail"""
        npsn = school_info['npsn']
        url = school_info['url']
        
        detail = self.scrape_school_detail(npsn, url)
        if detail:
            # merge dengan info dasar
            detail.update({
                'nama_sekolah': school_info.get('nama_sekolah', detail['nama_sekolah']),
                'status_sekolah': school_info.get('status_sekolah', detail['status_sekolah']),
                'alamat_jalan': school_info.get('alamat_jalan', detail['alamat_sekolah'])
            })
        return detail

    def scrape_all(self, max_pages=None):
        """Fungsi utama scraping"""
        start_time = time.time()
        
        # load checkpoint jika ada
        checkpoint = self.load_checkpoint()
        start_page = 1
        
        if checkpoint:
            resume = input(f"\nLanjutkan dari halaman {checkpoint['last_page']}? (y/n): ").strip().lower()
            if resume == 'y':
                start_page = checkpoint['last_page']
                self.all_data = self.load_temp_data()
                print(f"✓ Dilanjutkan dengan {len(self.all_data)} data yang sudah ada\n")

        print("→ Memulai proses scraping...\n")
        
        page = start_page
        
        try:
            while True:
                if max_pages and page > max_pages:
                    break

                print(f"[HALAMAN {page}] Scraping...")
                
                # ambil school list dari halaman
                schools = self.scrape_page(page)
                
                if not schools:
                    print(f"  ✗ Tidak ada sekolah ditemukan, berhenti")
                    break

                print(f"  → Ditemukan {len(schools)} sekolah, mengambil detail...")

                # proses setiap sekolah dengan threading
                with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                    futures = {executor.submit(self.process_school, school): school for school in schools}
                    
                    for future in as_completed(futures):
                        try:
                            detail = future.result()
                            if detail:
                                with self.lock:
                                    self.all_data.append(detail)
                                    
                                    # simpan batch
                                    if len(self.all_data) % self.batch_size == 0:
                                        self.save_batch()
                                        
                        except Exception as e:
                            if self.debug:
                                print(f"    ⚠ Error future: {e}")

                rate = len(self.all_data) / (time.time() - start_time)
                print(f"  ✓ Halaman {page} selesai | Total: {len(self.all_data):,} | Kecepatan: {rate:.1f}/s\n")

                # simpan checkpoint setiap halaman
                self.save_checkpoint(page, len(self.all_data), 0)
                self.save_temp_data(self.all_data)

                page += 1
                time.sleep(2)

        except KeyboardInterrupt:
            print("\n\n[INTERRUPT] Dihentikan oleh user")

        # simpan data final
        if self.all_data:
            filename = self.save_to_excel(self.all_data)
            
            elapsed = time.time() - start_time
            print(f"\n{'='*70}")
            print(f"[SELESAI]")
            print(f"  Total sekolah: {len(self.all_data):,}")
            print(f"  Waktu: {elapsed/60:.2f} menit")
            print(f"  Kecepatan: {len(self.all_data)/elapsed:.1f} sekolah/detik")
            print(f"  File: {filename}")
            print(f"{'='*70}\n")

            # cleanup
            if os.path.exists(self.checkpoint_file):
                os.remove(self.checkpoint_file)
            if os.path.exists(self.temp_data_file):
                os.remove(self.temp_data_file)
        else:
            print("\n[ERROR] Tidak ada data terkumpul!")

        return self.all_data

    def save_batch(self):
        """Simpan data batch"""
        try:
            start_idx = (self.batch_counter * self.batch_size) + 1
            end_idx = start_idx + self.batch_size - 1
            
            batch_data = self.all_data[start_idx-1:end_idx]
            
            if batch_data:
                filename = f"batch_{start_idx}-{end_idx}.csv"
                df = pd.DataFrame(batch_data)
                df.to_csv(filename, index=False, encoding='utf-8-sig')
                print(f"  Batch tersimpan: {filename}")
                
                self.batch_counter += 1
                
        except Exception as e:
            if self.debug:
                print(f"  Error simpan batch: {e}")

    def save_to_excel(self, data):
        """Simpan ke Excel dengan formatting"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data Sekolah"

            # ambil semua unique keys
            all_keys = set()
            for item in data:
                all_keys.update(item.keys())

            headers = sorted(list(all_keys))
            ws.append(headers)

            # format header
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # tambah data
            for school in data:
                row = [school.get(key, '-') for key in headers]
                ws.append(row)

            # auto-adjust lebar kolom
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 100)

            filename = f"data_sekolah_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)

            return filename

        except Exception as e:
            print(f"[ERROR] Gagal simpan Excel: {e}")
            
            # fallback ke CSV
            try:
                filename = f"data_sekolah_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                df = pd.DataFrame(data)
                df.to_csv(filename, index=False, encoding='utf-8-sig')
                return filename
            except Exception as e2:
                print(f"[ERROR] CSV fallback gagal: {e2}")
                return None

    def save_checkpoint(self, page, count, total):
        """Simpan checkpoint"""
        with self.lock:
            with open(self.checkpoint_file, 'w', encoding='utf-8') as f:
                json.dump({
                    'last_page': page,
                    'processed_count': count,
                    'total_count': total,
                    'timestamp': datetime.now().isoformat()
                }, f, indent=2)

    def load_checkpoint(self):
        """Muat checkpoint"""
        if os.path.exists(self.checkpoint_file):
            try:
                with open(self.checkpoint_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                return None
        return None

    def save_temp_data(self, data):
        """Simpan temporary data"""
        with self.lock:
            try:
                with open(self.temp_data_file, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=2, ensure_ascii=False)
            except Exception as e:
                if self.debug:
                    print(f"  Error simpan temp: {e}")

    def load_temp_data(self):
        """Muat temporary data"""
        if os.path.exists(self.temp_data_file):
            try:
                with open(self.temp_data_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                return []
        return []


if __name__ == "__main__":
    print("\n" + "="*70)
    print("  SEKOLAH SCRAPER - SELENIUM COMPLETE VERSION")
    print("="*70)

    test_mode = input("\nMode test (scrape 3 halaman saja)? (y/n): ").strip().lower()
    max_pages = 3 if test_mode == 'y' else None

    workers = input("Jumlah parallel workers (default=2): ").strip()
    workers = int(workers) if workers.isdigit() else 2

    headless = input("Gunakan headless browser? (y/n, default=y): ").strip().lower()
    headless = headless != 'n'

    debug = input("Aktifkan debug mode? (y/n, default=n): ").strip().lower()
    debug = debug == 'y'

    print("\n" + "="*70)
    print("MEMULAI SCRAPER...")
    print("="*70 + "\n")

    try:
        scraper = SekolahScraperSelenium(
            max_workers=workers,
            headless=headless,
            debug=debug
        )

        data = scraper.scrape_all(max_pages=max_pages)

    except KeyboardInterrupt:
        print("\n\n[EXIT] Program dihentikan oleh user")
    except Exception as e:
        print(f"\n[ERROR] {e}")
        import traceback
        traceback.print_exc()