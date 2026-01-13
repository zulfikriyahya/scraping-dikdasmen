"""
SEKOLAH SCRAPER - SELENIUM COMPLETE VERSION
============================================
Scraping data sekolah dari sekolah.data.kemendikdasmen.go.id
Menggunakan Selenium dengan headless browser
"""

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from bs4 import BeautifulSoup
import pandas as pd
import time
import json
import os
import re
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


class SekolahScraperSelenium:
    """
    Selenium-based scraper untuk data sekolah
    """

    def __init__(self, max_workers=2, headless=True, debug=False):
        """
        Parameters:
        - max_workers: jumlah browser instances untuk parallel scraping
        - headless: True untuk headless browser
        - debug: True untuk verbose logging
        """
        self.max_workers = max_workers
        self.headless = headless
        self.debug = debug

        self.base_url = "https://sekolah.data.kemendikdasmen.go.id"
        self.checkpoint_file = "checkpoint.json"
        self.temp_data_file = "temp_data.json"
        self.batch_size = 50
        self.batch_counter = 0

        self.lock = threading.Lock()
        self.all_data = []

        print("\n" + "="*70)
        print("  SEKOLAH SCRAPER - SELENIUM COMPLETE VERSION")
        print("="*70)
        print(f"  Workers: {max_workers}")
        print(f"  Headless: {headless}")
        print(f"  Debug: {debug}")
        print("="*70 + "\n")

    def create_driver(self):
        """Membuat instance Selenium WebDriver"""
        try:
            chrome_options = Options()

            if self.headless:
                chrome_options.add_argument('--headless=new')

            chrome_options.add_argument('--no-sandbox')
            chrome_options.add_argument('--disable-dev-shm-usage')
            chrome_options.add_argument('--disable-blink-features=AutomationControlled')
            chrome_options.add_argument('--disable-gpu')
            chrome_options.add_argument('--window-size=1920,1080')
            chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')

            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)

            driver = webdriver.Chrome(options=chrome_options)
            driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

            return driver

        except Exception as e:
            print(f"[ERROR] Gagal membuat driver: {e}")
            return None

    def get_value_by_label(self, soup, label_text):
        """Helper untuk mengambil value berdasarkan label"""
        try:
            # Cari elemen yang mengandung label
            label_elem = soup.find(string=lambda x: x and label_text in str(x))
            if label_elem:
                parent = label_elem.find_parent()
                if parent:
                    # Cari sibling atau child berikutnya dengan class font-semibold
                    value_elem = parent.find_next('div', class_=lambda x: x and 'font-semibold' in str(x))
                    if value_elem:
                        text = value_elem.get_text(strip=True)
                        # Filter simbol kosong (em dash, en dash, minus, hyphen, dll)
                        empty_symbols = ['—', '–', '-', '', ' ', '−', '―', '‐', '‑', '‒', '–', '—', '―']
                        if text and text not in empty_symbols:
                            return text
        except:
            pass
        return '-'

    def extract_school_list(self, driver):
        """Ekstrak daftar sekolah dari halaman list"""
        try:
            time.sleep(3)
            
            # Gunakan Selenium untuk klik dan ambil URL
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            
            schools = []
            
            # Cari semua artikel dengan Selenium
            article_elements = driver.find_elements(By.TAG_NAME, "article")
            
            for idx, article_elem in enumerate(article_elements):
                try:
                    # Parse dengan BeautifulSoup untuk data text
                    article_html = article_elem.get_attribute('outerHTML')
                    article_soup = BeautifulSoup(article_html, 'html.parser')
                    
                    # Extract NPSN dengan regex
                    npsn = ''
                    npsn_text = article_soup.find(string=lambda x: x and 'NPSN' in str(x))
                    if npsn_text:
                        match = re.search(r'NPSN\s*:\s*(\w+)', str(npsn_text))
                        if match:
                            npsn = match.group(1).strip()
                    
                    # Extract nama sekolah dari h3
                    nama = ''
                    nama_elem = article_soup.find('h3')
                    if nama_elem:
                        nama = nama_elem.get_text(strip=True)
                    
                    # Extract status (Negeri/Swasta)
                    status = ''
                    status_elem = article_soup.find('div', class_=lambda x: x and 'text-orange-600' in str(x))
                    if not status_elem:
                        status_elem = article_soup.find('div', class_=lambda x: x and 'font-semibold' in str(x))
                    if status_elem:
                        status = status_elem.get_text(strip=True)
                    
                    # Extract alamat
                    alamat = ''
                    alamat_elem = article_soup.find('div', class_=lambda x: x and 'line-clamp' in str(x) and 'text-xs' in str(x))
                    if alamat_elem:
                        alamat = alamat_elem.get_text(strip=True)
                    
                    # Extract URL dengan cara klik tombol Lihat dan ambil URL
                    url = ''
                    try:
                        # Cari tombol "Lihat" dengan selector yang spesifik
                        # Method 1: Cari button dengan span yang textnya "Lihat"
                        try:
                            button = article_elem.find_element(By.XPATH, ".//button[.//span[text()='Lihat']]")
                        except:
                            # Method 2: Cari button dengan label="Lihat"
                            try:
                                button = article_elem.find_element(By.XPATH, ".//button[@label='Lihat']")
                            except:
                                # Method 3: Cari button dengan class p-button-primary yang ada text Lihat
                                button = article_elem.find_element(By.XPATH, ".//button[contains(@class, 'p-button-primary')]")
                        
                        # Simpan URL saat ini
                        current_url = driver.current_url
                        
                        # Klik tombol dengan JavaScript untuk memastikan klik berhasil
                        driver.execute_script("arguments[0].scrollIntoView(true);", button)
                        time.sleep(0.5)
                        driver.execute_script("arguments[0].click();", button)
                        time.sleep(2)
                        
                        # Ambil URL baru (profil sekolah)
                        new_url = driver.current_url
                        
                        if new_url != current_url and 'profil-sekolah' in new_url:
                            url = new_url
                            
                            # Kembali ke halaman list
                            driver.back()
                            time.sleep(2)
                        else:
                            # Jika tidak redirect, coba cari dari onclick atau href
                            url = ''
                    except Exception as e:
                        # Fallback: coba cari URL dari routerLink atau attribute lain
                        if self.debug:
                            print(f"      ⚠ Gagal klik tombol Lihat untuk artikel {idx}: {e}")
                        pass
                    
                    if npsn and url:
                        schools.append({
                            'npsn': npsn,
                            'url': url,
                            'nama_sekolah': nama,
                            'status_sekolah': status,
                            'alamat_jalan': alamat
                        })
                    elif npsn:
                        # Jika URL tidak didapat, skip atau gunakan placeholder
                        if self.debug:
                            print(f"    ⚠ URL tidak ditemukan untuk NPSN {npsn}")
                
                except Exception as e:
                    if self.debug:
                        print(f"    ⚠ Parse article {idx} error: {e}")
                    continue
            
            return schools
        
        except Exception as e:
            print(f"  ✗ Extract list error: {e}")
            return []

    def scrape_school_detail(self, url, basic_info):
        """Scrape detail sekolah dari halaman profil"""
        driver = self.create_driver()
        if not driver:
            return None

        try:
            driver.get(url)
            time.sleep(4)

            soup = BeautifulSoup(driver.page_source, 'html.parser')

            data = {
                'npsn': basic_info.get('npsn', '-'),
                'url': url,
                'nama_sekolah': basic_info.get('nama_sekolah', '-'),
                'alamat_sekolah': '-',
                'alamat_jalan': basic_info.get('alamat_jalan', '-'),
                'url_sk_operasional': '-',
                'akreditasi': '-',
                'kepala_sekolah': '-',
                'telepon': '-',
                'website': '-',
                'status_sekolah': basic_info.get('status_sekolah', '-'),
                'bentuk_pendidikan': '-',
                'operator': '-',
                'email': '-',
                'yayasan': '-',
                # Statistik
                'guru': '-',
                'siswa_laki': '-',
                'siswa_perempuan': '-',
                'rombongan_belajar': '-',
                'daya_tampung': '-',
                'ruang_kelas': '-',
                'laboratorium': '-',
                'perpustakaan': '-',
                # Kurikulum & Utilitas
                'kurikulum': '-',
                'penyelenggaraan': '-',
                'akses_internet': '-',
                'sumber_listrik': '-',
                'daya_listrik': '-',
                'luas_tanah': '-',
                # Proses Pembelajaran
                'rasio_siswa_rombel': '-',
                'rasio_rombel_ruang_kelas': '-',
                'rasio_siswa_guru': '-',
                'persen_guru_kualifikasi': '-',
                'persen_guru_sertifikasi': '-',
                'persen_guru_pns': '-',
                'persen_ruang_kelas_layak': '-',
                # Alamat
                'lintang': '-',
                'bujur': '-',
                'link_google_map': '-'
            }

            # Extract nama sekolah dari h1
            h1 = soup.find('h1')
            if h1:
                data['nama_sekolah'] = h1.get_text(strip=True)

            # Extract alamat sekolah dari p setelah h1
            p_alamat = soup.find('p', class_=lambda x: x and 'text-slate-600' in str(x))
            if p_alamat:
                data['alamat_sekolah'] = p_alamat.get_text(strip=True)

            # Extract SK Operasional
            # Method 1: Construct dari UUID di URL
            try:
                # Ambil UUID dari URL profil
                uuid_match = re.search(r'/profil-sekolah/([A-F0-9-]+)', url, re.IGNORECASE)
                if uuid_match:
                    uuid = uuid_match.group(1)
                    # Construct URL SK Operasional
                    sk_url = f"https://file.data.kemendikdasmen.go.id/sk/{uuid}.pdf"
                    data['url_sk_operasional'] = sk_url
                else:
                    data['url_sk_operasional'] = '-'
            except:
                data['url_sk_operasional'] = '-'
            
            # Method 2: Jika method 1 gagal, coba klik tombol Aksi (optional, bisa di-skip karena lebih lambat)
            if data['url_sk_operasional'] == '-':
                try:
                    aksi_button = driver.find_element(By.XPATH, "//button[.//span[text()='Aksi']]")
                    if aksi_button:
                        driver.execute_script("arguments[0].click();", aksi_button)
                        time.sleep(1)
                        
                        soup_after_click = BeautifulSoup(driver.page_source, 'html.parser')
                        sk_menu = soup_after_click.find('span', class_='p-menu-item-label', 
                                                       string=lambda x: x and 'SK Operasional' in str(x))
                        
                        if sk_menu:
                            parent_a = sk_menu.find_parent('a')
                            if parent_a and parent_a.get('href'):
                                data['url_sk_operasional'] = parent_a.get('href')
                        
                        # Tutup menu
                        driver.execute_script("document.body.click();")
                        time.sleep(0.5)
                except:
                    pass

            # Extract data dari grid profil umum menggunakan helper
            data['akreditasi'] = self.get_value_by_label(soup, 'Akreditasi')
            data['status_sekolah'] = self.get_value_by_label(soup, 'Status Sekolah') or data['status_sekolah']
            
            # NPSN - cari link ke referensi.data.kemendikdasmen.go.id
            npsn_link = soup.find('a', href=lambda x: x and 'referensi.data.kemendikdasmen.go.id' in str(x) and 'npsn' in str(x))
            if npsn_link:
                npsn_text = npsn_link.get_text(strip=True)
                # Bersihkan dari whitespace dan icon
                npsn_text = npsn_text.strip()
                if npsn_text and (npsn_text.startswith('P') or npsn_text.startswith('p')):
                    data['npsn'] = npsn_text
                else:
                    data['npsn'] = basic_info.get('npsn', '-')
            else:
                # Fallback: coba dari label atau basic_info
                npsn_val = self.get_value_by_label(soup, 'NPSN')
                if npsn_val and npsn_val != '-' and (npsn_val.startswith('P') or npsn_val.startswith('p')):
                    data['npsn'] = npsn_val
                else:
                    data['npsn'] = basic_info.get('npsn', '-')
            
            data['bentuk_pendidikan'] = self.get_value_by_label(soup, 'Bentuk Pendidikan')
            data['kepala_sekolah'] = self.get_value_by_label(soup, 'Kepala Sekolah')
            data['operator'] = self.get_value_by_label(soup, 'Operator')
            
            # Telepon - cari icon pi-phone
            phone_section = soup.find('i', class_=lambda x: x and 'pi-phone' in str(x))
            if phone_section:
                phone_parent = phone_section.find_parent()
                if phone_parent:
                    # Coba cari link tel:
                    phone_link = phone_parent.find('a', href=lambda x: x and 'tel:' in str(x))
                    if phone_link:
                        data['telepon'] = phone_link.get_text(strip=True)
                    else:
                        # Cari span atau div dengan font-semibold
                        phone_elem = phone_parent.find(['span', 'div'], class_=lambda x: x and 'font-semibold' in str(x))
                        if phone_elem:
                            phone_text = phone_elem.get_text(strip=True)
                            # Filter simbol kosong yang lengkap
                            if phone_text and phone_text not in ['—', '–', '-', '', ' ', '−', '―']:
                                data['telepon'] = phone_text

            # Email - cari link mailto
            # Method 1: Cari dari icon pi-envelope
            email_section = soup.find('i', class_=lambda x: x and 'pi-envelope' in str(x))
            if email_section:
                email_parent = email_section.find_parent()
                # Cari beberapa level ke atas jika perlu
                for _ in range(3):  # Maksimal 3 level ke atas
                    if email_parent:
                        email_link = email_parent.find('a', href=lambda x: x and 'mailto:' in str(x))
                        if email_link:
                            data['email'] = email_link.get_text(strip=True)
                            break
                        email_parent = email_parent.find_parent()
            
            # Method 2: Fallback - cari semua link mailto di halaman
            if data['email'] == '-':
                all_mailto = soup.find_all('a', href=lambda x: x and 'mailto:' in str(x))
                # Ambil yang pertama (biasanya email sekolah)
                if all_mailto:
                    # Filter yang bukan email umum/support
                    for mailto in all_mailto:
                        email_text = mailto.get_text(strip=True)
                        # Pastikan ada @ dan bukan email support/admin
                        if '@' in email_text and 'support' not in email_text.lower():
                            data['email'] = email_text
                            break

            # Website - cari icon pi-globe
            web_section = soup.find('i', class_=lambda x: x and 'pi-globe' in str(x))
            if web_section:
                web_parent = web_section.find_parent()
                if web_parent:
                    # Cari link dengan http
                    web_link = web_parent.find('a', href=lambda x: x and 'http' in str(x))
                    if web_link:
                        data['website'] = web_link.get('href', '-')
                    else:
                        # Cari div dengan font-semibold
                        web_elem = web_parent.find('div', class_=lambda x: x and 'font-semibold' in str(x))
                        if web_elem:
                            web_text = web_elem.get_text(strip=True)
                            # Filter text kosong atau placeholder
                            invalid_texts = ['—', '–', '-', '', ' ', '−', '―',
                                           'Belum Terisi', 'Belum terisi', 'belum terisi', 
                                           'BELUM TERISI', 'Tidak Ada', 'tidak ada',
                                           'Tidak ada', 'TIDAK ADA']
                            if web_text and web_text not in invalid_texts:
                                data['website'] = web_text

            data['yayasan'] = self.get_value_by_label(soup, 'Yayasan')

            # Extract statistik sekolah - cari card dengan h2 "Statistik Sekolah"
            stat_section = soup.find('h2', string=lambda x: x and 'Statistik Sekolah' in str(x))
            if stat_section:
                stat_parent = stat_section.find_parent()
                if stat_parent:
                    # Cari semua item statistik
                    stat_items = stat_parent.find_all('div', class_=lambda x: x and 'text-2xl' in str(x))
                    stat_labels = stat_parent.find_all('div', class_=lambda x: x and 'text-slate-600' in str(x))

                    for stat_val, stat_lbl in zip(stat_items, stat_labels):
                        label = stat_lbl.get_text(strip=True).lower()
                        value = stat_val.get_text(strip=True)

                        if 'guru' in label and 'kualifikasi' not in label:
                            data['guru'] = value
                        elif 'laki-laki' in label or 'laki laki' in label:
                            data['siswa_laki'] = value
                        elif 'perempuan' in label:
                            data['siswa_perempuan'] = value
                        elif 'rombongan' in label or 'rombel' in label:
                            data['rombongan_belajar'] = value
                        elif 'daya tampung' in label:
                            data['daya_tampung'] = value
                        elif 'ruang kelas' in label:
                            data['ruang_kelas'] = value
                        elif 'laboratorium' in label:
                            data['laboratorium'] = value
                        elif 'perpustakaan' in label:
                            data['perpustakaan'] = value

            # Extract kurikulum & utilitas
            kurikulum_section = soup.find('h2', string=lambda x: x and 'Kurikulum' in str(x))
            if kurikulum_section:
                kurikulum_parent = kurikulum_section.find_parent()
                if kurikulum_parent:
                    data['kurikulum'] = self.get_value_by_label(kurikulum_parent, 'Kurikulum')
                    data['penyelenggaraan'] = self.get_value_by_label(kurikulum_parent, 'Penyelenggaraan')
                    data['akses_internet'] = self.get_value_by_label(kurikulum_parent, 'Akses Internet')
                    data['sumber_listrik'] = self.get_value_by_label(kurikulum_parent, 'Sumber Listrik')
                    data['daya_listrik'] = self.get_value_by_label(kurikulum_parent, 'Daya Listrik')
                    data['luas_tanah'] = self.get_value_by_label(kurikulum_parent, 'Luas Tanah')

            # Extract proses pembelajaran
            pembelajaran_section = soup.find('h2', string=lambda x: x and 'Proses Pembelajaran' in str(x))
            if pembelajaran_section:
                pembelajaran_parent = pembelajaran_section.find_parent()
                if pembelajaran_parent:
                    # Cari semua item dengan struktur khusus
                    items = pembelajaran_parent.find_all('li')
                    for item in items:
                        label_elem = item.find('div', class_=lambda x: x and 'text-slate-600' in str(x))
                        value_elem = item.find('div', class_=lambda x: x and 'font-semibold' in str(x))
                        
                        if label_elem and value_elem:
                            label = label_elem.get_text(strip=True).lower()
                            value = value_elem.get_text(strip=True)

                            if 'rasio siswa rombel' in label:
                                data['rasio_siswa_rombel'] = value
                            elif 'rasio rombel ruang kelas' in label:
                                data['rasio_rombel_ruang_kelas'] = value
                            elif 'rasio siswa guru' in label:
                                data['rasio_siswa_guru'] = value
                            elif 'guru kualifikasi' in label:
                                data['persen_guru_kualifikasi'] = value
                            elif 'guru sertifikasi' in label:
                                data['persen_guru_sertifikasi'] = value
                            elif 'guru pns' in label:
                                data['persen_guru_pns'] = value
                            elif 'ruang kelas layak' in label:
                                data['persen_ruang_kelas_layak'] = value

            # Extract alamat dan koordinat
            alamat_section = soup.find('h2', string=lambda x: x and 'Alamat' in str(x))
            if alamat_section:
                alamat_parent = alamat_section.find_parent()
                if alamat_parent:
                    # Extract koordinat dari iframe Google Maps
                    iframe = alamat_parent.find('iframe')
                    if iframe and iframe.get('src'):
                        src = iframe.get('src')
                        # Parse koordinat dari URL maps
                        coord_match = re.search(r'q=([-\d.]+),([-\d.]+)', src)
                        if coord_match:
                            data['lintang'] = coord_match.group(1)
                            data['bujur'] = coord_match.group(2)
                            data['link_google_map'] = f"https://www.google.com/maps?q={coord_match.group(1)},{coord_match.group(2)}"

                    # Extract koordinat dari text (backup)
                    lat_elem = alamat_parent.find('div', string=lambda x: x and 'Lintang' in str(x))
                    if lat_elem:
                        lat_parent = lat_elem.find_parent()
                        if lat_parent:
                            lat_val = lat_parent.find_next('div', class_=lambda x: x and 'font-medium' in str(x))
                            if lat_val:
                                data['lintang'] = lat_val.get_text(strip=True)

                    lon_elem = alamat_parent.find('div', string=lambda x: x and 'Bujur' in str(x))
                    if lon_elem:
                        lon_parent = lon_elem.find_parent()
                        if lon_parent:
                            lon_val = lon_parent.find_next('div', class_=lambda x: x and 'font-medium' in str(x))
                            if lon_val:
                                data['bujur'] = lon_val.get_text(strip=True)

                    # Cari tombol Google Maps atau Salin Tautan
                    map_buttons = alamat_parent.find_all('button')
                    for btn in map_buttons:
                        btn_text = btn.get_text(strip=True).lower()
                        if 'google maps' in btn_text or 'salin' in btn_text:
                            # Tombol biasanya trigger action JS, koordinat sudah diambil dari iframe
                            pass

            return data

        except Exception as e:
            if self.debug:
                print(f"  ✗ Detail scrape error untuk {npsn}: {e}")
            return None
        finally:
            driver.quit()

    def scrape_page(self, page_num):
        """Scrape satu halaman list"""
        driver = self.create_driver()
        if not driver:
            return []

        try:
            url = f"{self.base_url}/sekolah"
            driver.get(url)
            time.sleep(5)

            # Navigate ke halaman yang diminta
            for i in range(1, page_num):
                try:
                    next_btn = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "button.p-paginator-next:not(.p-disabled)"))
                    )
                    next_btn.click()
                    time.sleep(3)
                except:
                    print(f"  ✗ Tidak bisa navigasi ke halaman {page_num}")
                    return []

            # Extract school list
            schools = self.extract_school_list(driver)

            return schools

        except Exception as e:
            print(f"  ✗ Page {page_num} error: {e}")
            return []
        finally:
            driver.quit()

    def process_school(self, school_info):
        """Proses detail satu sekolah"""
        url = school_info.get('url')
        
        if not url:
            if self.debug:
                print(f"    ⚠ URL tidak tersedia untuk {school_info.get('npsn', 'unknown')}")
            return None
        
        detail = self.scrape_school_detail(url, school_info)
        
        return detail

    def scrape_all(self, max_pages=None):
        """Fungsi utama scraping"""
        start_time = time.time()

        # Load checkpoint jika ada
        checkpoint = self.load_checkpoint()
        start_page = 1

        if checkpoint:
            resume = input(f"\nResume dari halaman {checkpoint['last_page']}? (y/n): ").strip().lower()
            if resume == 'y':
                start_page = checkpoint['last_page']
                self.all_data = self.load_temp_data()
                print(f"✓ Resume dengan {len(self.all_data)} data existing\n")

        print("→ Memulai proses scraping...\n")

        page = start_page

        try:
            while True:
                if max_pages and page > max_pages:
                    break

                print(f"[HALAMAN {page}] Scraping...")

                # Ambil daftar sekolah dari halaman
                schools = self.scrape_page(page)

                if not schools:
                    print(f"  ✗ Tidak ada sekolah ditemukan, berhenti")
                    break

                print(f"  → Ditemukan {len(schools)} sekolah, ambil detail...")

                # Proses setiap sekolah dengan threading
                with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                    futures = {executor.submit(self.process_school, school): school for school in schools}

                    for future in as_completed(futures):
                        try:
                            detail = future.result()
                            if detail:
                                with self.lock:
                                    self.all_data.append(detail)

                                    # Simpan batch
                                    if len(self.all_data) % self.batch_size == 0:
                                        self.save_batch()

                        except Exception as e:
                            if self.debug:
                                print(f"    ⚠ Future error: {e}")

                rate = len(self.all_data) / (time.time() - start_time)
                print(f"  ✓ Halaman {page} selesai | Total: {len(self.all_data):,} | Rate: {rate:.1f}/s\n")

                # Simpan checkpoint setiap halaman
                self.save_checkpoint(page, len(self.all_data), 0)
                self.save_temp_data(self.all_data)

                page += 1
                time.sleep(2)

        except KeyboardInterrupt:
            print("\n\n[INTERRUPT] Dihentikan oleh user")

        # Simpan data final
        if self.all_data:
            filename = self.save_to_excel(self.all_data)

            elapsed = time.time() - start_time
            print(f"\n{'='*70}")
            print(f"[SELESAI]")
            print(f"  Total sekolah: {len(self.all_data):,}")
            print(f"  Waktu: {elapsed/60:.2f} menit")
            print(f"  Rate: {len(self.all_data)/elapsed:.1f} sekolah/detik")
            print(f"  File: {filename}")
            print(f"{'='*70}\n")

            # Cleanup
            if os.path.exists(self.checkpoint_file):
                os.remove(self.checkpoint_file)
            if os.path.exists(self.temp_data_file):
                os.remove(self.temp_data_file)
        else:
            print("\n[ERROR] Tidak ada data terkumpul!")

        return self.all_data

    def save_batch(self):
        """Simpan batch data"""
        try:
            start_idx = (self.batch_counter * self.batch_size) + 1
            end_idx = start_idx + self.batch_size - 1

            batch_data = self.all_data[start_idx-1:end_idx]

            if batch_data:
                filename = f"batch_{start_idx}-{end_idx}.csv"
                df = pd.DataFrame(batch_data)
                df.to_csv(filename, index=False, encoding='utf-8-sig')
                print(f"  💾 Batch disimpan: {filename}")

                self.batch_counter += 1

        except Exception as e:
            if self.debug:
                print(f"  ⚠ Batch save error: {e}")

    def save_to_excel(self, data):
        """Simpan ke Excel dengan formatting"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data Sekolah"

            # Header sesuai urutan dari requirements
            headers = [
                'npsn', 'url', 'nama_sekolah', 'alamat_sekolah', 'alamat_jalan',
                'url_sk_operasional', 'akreditasi', 'kepala_sekolah', 'telepon',
                'website', 'status_sekolah', 'bentuk_pendidikan', 'operator',
                'email', 'yayasan', 'guru', 'siswa_laki', 'siswa_perempuan',
                'rombongan_belajar', 'daya_tampung', 'ruang_kelas', 'laboratorium',
                'perpustakaan', 'kurikulum', 'penyelenggaraan', 'akses_internet',
                'sumber_listrik', 'daya_listrik', 'luas_tanah', 'rasio_siswa_rombel',
                'rasio_rombel_ruang_kelas', 'rasio_siswa_guru', 'persen_guru_kualifikasi',
                'persen_guru_sertifikasi', 'persen_guru_pns', 'persen_ruang_kelas_layak',
                'lintang', 'bujur', 'link_google_map'
            ]
            
            ws.append(headers)

            # Format header
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Tambah data
            for school in data:
                row = [school.get(key, '-') for key in headers]
                ws.append(row)

            # Auto-adjust lebar kolom
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 100)

            filename = f"data_sekolah_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)

            return filename

        except Exception as e:
            print(f"[ERROR] Gagal simpan Excel: {e}")

            # Fallback ke CSV
            try:
                filename = f"data_sekolah_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                df = pd.DataFrame(data)
                df.to_csv(filename, index=False, encoding='utf-8-sig')
                return filename
            except Exception as e2:
                print(f"[ERROR] CSV fallback gagal: {e2}")
                return None

    def save_checkpoint(self, page, count, total):
        """Simpan checkpoint"""
        with self.lock:
            with open(self.checkpoint_file, 'w', encoding='utf-8') as f:
                json.dump({
                    'last_page': page,
                    'processed_count': count,
                    'total_count': total,
                    'timestamp': datetime.now().isoformat()
                }, f, indent=2)

    def load_checkpoint(self):
        """Load checkpoint"""
        if os.path.exists(self.checkpoint_file):
            try:
                with open(self.checkpoint_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                return None
        return None

    def save_temp_data(self, data):
        """Simpan data sementara"""
        with self.lock:
            try:
                with open(self.temp_data_file, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=2, ensure_ascii=False)
            except Exception as e:
                if self.debug:
                    print(f"  ⚠ Temp save error: {e}")

    def load_temp_data(self):
        """Load data sementara"""
        if os.path.exists(self.temp_data_file):
            try:
                with open(self.temp_data_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                return []
        return []


if __name__ == "__main__":
    print("\n" + "="*70)
    print("  SEKOLAH SCRAPER - SELENIUM COMPLETE VERSION")
    print("="*70)

    test_mode = input("\nMode test (scrape 3 halaman saja)? (y/n): ").strip().lower()
    max_pages = 3 if test_mode == 'y' else None

    workers = input("Jumlah parallel workers (default=2): ").strip()
    workers = int(workers) if workers.isdigit() else 2

    headless = input("Gunakan headless browser? (y/n, default=y): ").strip().lower()
    headless = headless != 'n'

    debug = input("Aktifkan debug mode? (y/n, default=n): ").strip().lower()
    debug = debug == 'y'

    print("\n" + "="*70)
    print("MEMULAI SCRAPER...")
    print("="*70 + "\n")

    try:
        scraper = SekolahScraperSelenium(
            max_workers=workers,
            headless=headless,
            debug=debug
        )

        data = scraper.scrape_all(max_pages=max_pages)

    except KeyboardInterrupt:
        print("\n\n[EXIT] Program dihentikan oleh user")
    except Exception as e:
        print(f"\n[ERROR] {e}")
        import traceback
        traceback.print_exc()